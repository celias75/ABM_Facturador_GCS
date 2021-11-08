#!/usr/bin/python
# -*- coding: latin-1 -*-
# ABM de usuarios, productos y productos en facturador GCS
# main -h para ayuda

import sys
import psycopg2
from psycopg2.extensions import AsIs
from openpyxl import load_workbook
import logging
import os
import argparse
import mapping


# PARAMETROS

log_file = 'abm.log'
DEBUG = True




def leer_hoja(sheet,config):
    '''
    Lee una hoja de excel y devuelve una lista de diccionarios
    :param sheet:
    :param config:
    :return:
    '''
    info = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {}
        for key in config:
            item[key] = row[config[key]]
        info.append(item)
    return info

def read_data(file_name):
    '''
    Lee un archivo excel y devuelve una lista de diccionarios
    :param file_name: archivo excel
    :return:
    '''
    usuarios =[]
    productos = []
    puntos_venta = []
    workbook = load_workbook(filename=file_name, read_only=True)
    for key in mapping.sheets:
        sheet = workbook[mapping.sheets[key]]
        logger.info ("Leyendo Hoja de " + key)
        if key == 'usuarios':
            usuarios = leer_hoja(sheet,mapping.usuario_xls)
            logger.debug(usuarios)
        elif key == 'productos':
            productos = leer_hoja(sheet,mapping.producto_xls)
            logger.debug(productos)
        elif key == 'puntos':
            puntos_venta = leer_hoja(sheet,mapping.puntos_xls)
            logger.debug(puntos_venta)
        else:
            logger.error("No se reconoce la hoja " + key)
            sys.exit(1)
    return usuarios, productos, puntos_venta

def cargarDB(items,cur,table):
    '''
    Carga una lista de diccionarios en una tabla de la base de datos
    :param items:
    :param cur:
    :param table:
    :return:
    '''

    for item in items:
        columns = item.keys()
        values = [item[column] for column in columns]
        insert_statement = f'INSERT INTO {table} (%s) VALUES %s'
        logger.debug(cur.mogrify(insert_statement, (AsIs(','.join(columns)), tuple(values))))
        cur.execute(insert_statement, (AsIs(','.join(columns)), tuple(values)))

    return

def cargarDB_avanzada(items_usuarios,cur,table1,table2):
    '''
    Carga la tabla usuarios y la tabla operadores con la relación usuario - punto de venta
    :param items:
    :param cur:
    :param table:
    :return:
    '''

    for usuario in items_usuarios:
        # Extraer de item los puntos de venta
        puntos_venta =[]
        operadores =[]
        temp = usuario['puntos_venta'].split(',')
        del usuario['puntos_venta']
        for i in temp:
            try:
                puntos_venta.append (int(i))
            except:
                logger.warning("Punto de venta no valido, se ignora: " + i)

        for i in puntos_venta:
            # Trae de la tabla el ID del punto de venta
            id_punto_venta = buscarIdPuntoVenta(i,cur)
            if id_punto_venta:
                registro_operador = {
                    'puntosventa_idpuntoventa': id_punto_venta
                }
            else:
                logger.warning("No se encontro el punto de venta: " + str(i))
                continue
            operadores.append(registro_operador)

        logger.info("Cargando usuario: " + usuario['nombre'])
        lista = ''
        for i in operadores:
            lista += str(i['puntosventa_idpuntoventa']) + ','
        logger.info(f'Con los siguientes id de punto de venta: {lista}')

        columns1 = usuario.keys()
        values1 = [usuario[column1] for column1 in columns1]
        insert_statement1 = f'INSERT INTO {table1} (%s) VALUES %s RETURNING id'
        logger.debug(cur.mogrify(insert_statement1, (AsIs(','.join(columns1)), tuple(values1))))
        cur.execute(insert_statement1, (AsIs(','.join(columns1)), tuple(values1)))
        user_id = cur.fetchone()[0]
        for operador in operadores:
            operador['usuario_id']= user_id
            columns2 = operador.keys()
            values2 = [operador[column2] for column2 in columns2]
            insert_statement2 = f'INSERT INTO {table2} (%s) VALUES %s'
            logger.debug(cur.mogrify(insert_statement2, (AsIs(','.join(columns2)), tuple(values2))))
            cur.execute(insert_statement2, (AsIs(','.join(columns2)), tuple(values2)))

    return

def buscarIdPuntoVenta(punto_venta,cur):
    '''
    Busca el ID de un punto de venta en la tabla puntos_venta
    :param id_punto_venta:
    :param cur:
    :return:
    '''
    query = f'SELECT idpuntoventa FROM {mapping.TABLA_PUNTOS} WHERE numero = {str(punto_venta)}'
    cur.execute(query)
    id_punto_venta = cur.fetchone()
    if id_punto_venta is None:
        logger.error("Punto de venta no encontrado")
        resultado = None
    else:
        resultado = id_punto_venta[0]

    return resultado


if __name__ == '__main__':

    # Inicializo log
    logger = logging.getLogger('abm')
    if DEBUG:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)
    # create file handler which logs even debug messages
    fh = logging.FileHandler(filename=log_file, encoding='UTF-8')
    # create console handler with a higher log level
    ch = logging.StreamHandler()
    # create formatter and add it to the handlers
    formatter_file = logging.Formatter('%(asctime)s %(levelname)-8s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    formatter_console = logging.Formatter('%(levelname)s - %(message)s')
    ch.setFormatter(formatter_console)
    fh.setFormatter(formatter_file)
    # add the handlers to logger
    logger.addHandler(ch)
    logger.addHandler(fh)

    # Inicializo parser de argumentos
    parser = argparse.ArgumentParser(description='ABM de usuarios, productos y puntos de venta en facturador GCS',
                                     epilog='Version 1.0 - (c) South Trade Network S.R.L 2021')
    parser.add_argument("ABM_file", nargs='?', help="Toma el archivo de de entrada especificado en "
                                                       "[ABM_file.xlsx]. Por defecto './Solicitud Facturador GCS.xlsx'",
                        default='./Solicitud Facturador GCS.xlsx')
    parser.add_argument('-u', '--usuarios', action='store_true', help='Carga de usuarios')
    parser.add_argument('-p', '--productos', action='store_true', help='Carga de productos')
    parser.add_argument('-s', '--puntos', action='store_true', help='Carga de puntos de venta')

    args = parser.parse_args()

    logger.info('-----------------------------------------------------------------------------------------------------')
    logger.info('Comienzo de Procesamiento')
    logger.info('Archivo de entrada: ' + args.ABM_file)
    param_puntos = param_usuarios = param_productos = 'FALSE'

    if args.puntos: param_puntos = 'TRUE'
    if args.usuarios: param_usuarios = 'TRUE'
    if args.productos: param_productos = 'TRUE'
    logger.info(f'Cargar puntos: {param_puntos}')
    logger.info(f'Cargar usuarios: {param_usuarios}')
    logger.info(f'Cargar productos: {param_productos}')
    
    # cargo archivo de entrada
    if not os.path.exists(args.ABM_file):
        logger.error(f'No se encuentra archivo de entrada {args.ABM_file}. Imposible continuar')
        sys.exit(1)
    try:
        usuarios, productos, puntos_venta = read_data(args.ABM_file)
    except Exception as e:
        logger.error(f'Error al leer archivo de entrada {args.ABM_file}. Imposible continuar')
        logger.error(e)
        sys.exit(1)

    # Inicializo conexion a DB

    try:
        import credentials
    except ImportError:
        logger.error('No se encuentra archivo credentials.py. Imposible continuar')
        sys.exit(1)

    try:
        conn = psycopg2.connect(**credentials.db)
        cur = conn.cursor()
        logger.info('Conexion a DB exitosa')
    except Exception as e:
        logger.error(f'Error al conectar a DB; {e}')
        sys.exit(1)

    # Cargo puntos de venta
    if args.puntos:
        if puntos_venta != []:
            # Validaciones y limpieza del array
            borrar = []
            for item in puntos_venta:
                if item['numero'] is None or item['descripcion'] is None or item['domicilio'] is None or \
                        item['numero'] =='' or item['descripcion'] == '' or item['domicilio'] == '':
                    borrar.append(item)
            if len(borrar) > 0:
                logger.warning(f'Se eliminan {len(borrar)} lineas de puntos de venta con datos incompletos')
                for i in borrar:
                    logger.warning(i)
                    puntos_venta.remove(i)
            # Cargo el array sanitizado
            if puntos_venta != []:
                # Vuelvo a preguntar porque quizá se me borraron todos
                logger.info('Cargando puntos de venta')
                cargarDB(puntos_venta,cur,mapping.TABLA_PUNTOS)
                conn.commit()
                logger.info(f'Puntos de Venta cargados: {len(puntos_venta)}')
        else:
            logger.warning('No hay puntos de venta para cargar')

    # Cargo productos
    if args.productos:
        if productos != []:
            # Validaciones y limpieza del array
            borrar = []
            for item in productos:
                if item['codigo'] is None or item['descripcion'] is None or item['precio'] is None  \
                        or item['unidadmedida_codigo']  is None or item['codigo'] =='' or item['descripcion'] =='':
                    borrar.append(item)
                elif item['nac_int'] != 1 and item['nac_int'] != 2:
                    borrar.append(item)
                elif not isinstance(item['precio'],int) and not isinstance(item['precio'],float):
                    borrar.append(item)
                elif not isinstance(item['unidadmedida_codigo'], int):
                    borrar.append(item)
            if len(borrar) > 0:
                logger.warning(f'Se eliminan {len(borrar)} lineas de productos con datos incompletos o incorrecto')
                for i in borrar:
                    logger.warning(i)
                    productos.remove(i)
            # Cargo el array sanitizado
            if productos != []:
                # Vuelvo a preguntar porque quizá se me borraron todos
                logger.info('Cargando productos')
                cargarDB(productos,cur,mapping.TABLA_PRODUCTOS)
                conn.commit()
                logger.info(f'Productos cargados: {len(productos)}')
        else:
            logger.warning('No hay productos para cargar')

    if args.usuarios:
        if usuarios != []:
            # Validaciones y limpieza del array
            borrar = []
            for item in usuarios:
                if item['email'] is None or item['puntos_venta'] is None or item['email'] == '' \
                        or item['puntos_venta'] == '':
                    borrar.append(item)
            if len(borrar) > 0:
                logger.warning(f'Se eliminan {len(borrar)} lineas de usuarios con datos incompletos')
                for i in borrar:
                    logger.warning(i)
                    usuarios.remove(i)

            # Cargo el array sanitizado
            if usuarios != []:
                # Vuelvo a preguntar porque quizá se me borraron todos
                logger.info('Cargando usuarios')
                # Agrego password a cada objeto dict
                for item in usuarios:
                    item['password'] = mapping.PASSWORD
                    item['activo'] = 1
                    item ['rol_id'] = 1
                cargarDB_avanzada(usuarios,cur,mapping.TABLA_USUARIOS,mapping.TABLA_OPERADORES)
                conn.commit()
                logger.info(f'Usuarios cargados: {len(usuarios)}')
        else:
            logger.warning('No hay usuarios para cargar')

    # cierro conexion a DB
    cur.close()
    conn.close()
    logger.info('Conexion a DB cerrada')

    logger.info('Carga de datos a DB exitosa')
    sys.exit(0)






